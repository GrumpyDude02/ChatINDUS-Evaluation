import os
import sqlite3
import openpyxl
from openpyxl.styles import Alignment,Border,Side


def extract_all_schemas_auto_detect(root_dir, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "All Schemas"

    sheet.append(["Database Name", "Table Name", "Column Name", "Data Type", "Number of Entries"])
    current_row = 2

    for folder in os.listdir(root_dir):
        folder_path = os.path.join(root_dir, folder)
        if not os.path.isdir(folder_path):
            continue

        db_file_path = None
        for file in os.listdir(folder_path):
            if file.endswith(".sqlite") or file.endswith(".db"):
                db_file_path = os.path.join(folder_path, file)
                break  # Take the first match

        if not db_file_path or not os.path.isfile(db_file_path):
            print(f"No database file found in: {folder}")
            continue

        try:
            conn = sqlite3.connect(db_file_path)
            cursor = conn.cursor()

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
            tables = [row[0] for row in cursor.fetchall()]
            db_start_row = current_row  # remember where this DB starts

            for table in tables:
                start_row = current_row

                cursor.execute(f"PRAGMA table_info('{table}')")
                columns = cursor.fetchall()

                cursor.execute(f"SELECT COUNT(*) FROM '{table}'")
                row_count = cursor.fetchone()[0]

                for col in columns:
                    col_name = col[1]
                    col_type = col[2]
                    sheet.append([folder, table, col_name, col_type, row_count])
                    current_row += 1

                # Merge "Table Name"
                if current_row - start_row > 1:
                    sheet.merge_cells(f"B{start_row}:B{current_row - 1}")
                    sheet.merge_cells(f"E{start_row}:E{current_row - 1}")
                    sheet[f"B{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
                    sheet[f"E{start_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Merge "Database Name" — after all its tables
            if current_row - db_start_row > 1:
                sheet.merge_cells(f"A{db_start_row}:A{current_row - 1}")
                sheet[f"A{db_start_row}"].alignment = Alignment(horizontal="center", vertical="center")

            conn.close()

        except Exception as e:
            print(f"Error processing {folder}: {e}")
    
    workbook.save(excel_file)
    print(f"All schemas saved to {excel_file}")

# Example usage
extract_all_schemas_auto_detect("databases", "all_schemas.xlsx")
